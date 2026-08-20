"""Escribir documentos: PDF y hoja de calculo.

Quien redacta es el motor; quien escribe el archivo es esto (regla 4). Aqui no
se le pregunta nada a ningun modelo: entra contenido, sale un archivo.

Dos cosas que no son detalle:

**El PDF lo pinta Qt, que ya esta instalado.** `QPdfWriter` + `QTextDocument`
maquetan markdown sencillo sin una sola dependencia nueva. Pero Qt **aborta el
proceso** —no lanza excepcion, aborta— si tocas `QTextDocument` sin un
`QGuiApplication` vivo. Y `friday.py --console` / `--say` no crean ninguno
(`run_console` es otro camino que `CompanionApp`). Sin el guardia de
`_hay_qt()`, pedir un PDF desde consola mataba a FRIDAY en seco, sin traza y
sin linea en la bitacora. Se comprueba y se dice, que es lo unico aceptable.

**`xlsx` es opcional.** openpyxl no viene de fabrica; sin el se escribe `.csv`,
que Excel abre igual. `formats()` dice la verdad de lo que hay instalado para
que la skill no prometa lo que no puede cumplir.
"""
from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any


def _hay_qt() -> bool:
    """¿Existe un QGuiApplication? Sin el, tocar Qt aborta el proceso."""
    try:
        from PySide6.QtGui import QGuiApplication
    except ImportError:
        return False
    return QGuiApplication.instance() is not None


def _hay_openpyxl() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


class LocalDocumentWriter:
    """Escribe PDF y hojas de calculo en disco. No lee nada del usuario."""

    def __init__(self, policy: Any = None):
        self.policy = policy

    # ── que se puede hacer aqui y ahora ───────────────────────────
    def formats(self) -> tuple[str, ...]:
        out = ["csv"]
        if _hay_openpyxl():
            out.append("xlsx")
        if _hay_qt():
            out.append("pdf")
        return tuple(out)

    # ── PDF ───────────────────────────────────────────────────────
    def write_pdf(self, path: Path, titulo: str, markdown: str) -> bool:
        if not _hay_qt():
            raise RuntimeError(
                "no puedo generar PDF sin la interfaz abierta "
                "(Qt necesita QGuiApplication; arranca sin --console)")

        from PySide6.QtCore import QMarginsF
        from PySide6.QtGui import QPageSize, QPdfWriter, QTextDocument

        path.parent.mkdir(parents=True, exist_ok=True)
        w = QPdfWriter(str(path))
        w.setPageSize(QPageSize(QPageSize.A4))
        w.setPageMargins(QMarginsF(18, 18, 18, 18))
        w.setTitle(titulo or path.stem)

        doc = QTextDocument()
        doc.setDefaultStyleSheet(
            "body{font-family:Segoe UI,sans-serif;font-size:11pt;color:#1a1a1a}"
            "h1{font-size:20pt;margin-bottom:2pt}"
            "h2{font-size:14pt;margin-top:14pt}"
            "td,th{padding:4px 8px;border:1px solid #bbb}"
            "code{font-family:Consolas,monospace;background:#f2f2f2}")
        doc.setHtml(f"<body>{_md_a_html(titulo, markdown)}</body>")
        doc.print_(w)
        return path.exists() and path.stat().st_size > 0

    # ── hoja de calculo ───────────────────────────────────────────
    def write_sheet(self, path: Path, cabeceras: list[str],
                    filas: list[list[Any]]) -> Path:
        """Devuelve la ruta REAL escrita: puede no ser la pedida.

        Si piden `.xlsx` y openpyxl no esta, se escribe `.csv` al lado. Se
        devuelve la ruta en vez de un bool para que la skill diga lo que de
        verdad hay en disco y no lo que se pidio.
        """
        path.parent.mkdir(parents=True, exist_ok=True)
        if path.suffix.lower() == ".xlsx" and _hay_openpyxl():
            from openpyxl import Workbook
            from openpyxl.styles import Font
            wb = Workbook()
            ws = wb.active
            ws.title = (path.stem[:31] or "Hoja")
            if cabeceras:
                ws.append(list(cabeceras))
                for c in ws[1]:
                    c.font = Font(bold=True)
            for fila in filas:
                ws.append(list(fila))
            for i, nombre in enumerate(cabeceras, start=1):
                largo = max([len(str(nombre))]
                            + [len(str(f[i - 1])) for f in filas if len(f) >= i] or [8])
                ws.column_dimensions[ws.cell(1, i).column_letter].width = min(largo + 2, 60)
            wb.save(str(path))
            return path

        destino = path.with_suffix(".csv")
        # `utf-8-sig`: sin BOM, Excel abre los acentos rotos en Windows.
        with destino.open("w", newline="", encoding="utf-8-sig") as fh:
            escritor = csv.writer(fh, delimiter=";")
            if cabeceras:
                escritor.writerow(cabeceras)
            escritor.writerows(filas)
        return destino


# ══════════════════════════════════════════════ markdown -> html
_LISTA = re.compile(r"^\s*[-*]\s+(.*)$")
_TITULO = re.compile(r"^(#{1,3})\s+(.*)$")


def _md_a_html(titulo: str, texto: str) -> str:
    """Markdown de andar por casa a HTML. No es un parser: es lo que el motor
    devuelve de verdad — titulos, listas, tablas y parrafos."""
    partes: list[str] = []
    if titulo:
        partes.append(f"<h1>{_esc(titulo)}</h1>")
    en_lista = False
    tabla: list[list[str]] = []

    def cerrar() -> None:
        nonlocal en_lista, tabla
        if en_lista:
            partes.append("</ul>")
            en_lista = False
        if tabla:
            partes.append(_tabla_html(tabla))
            tabla = []

    for linea in (texto or "").splitlines():
        cruda = linea.rstrip()
        if not cruda.strip():
            cerrar()
            continue
        if cruda.lstrip().startswith("|") and cruda.rstrip().endswith("|"):
            celdas = [c.strip() for c in cruda.strip().strip("|").split("|")]
            if not all(set(c) <= set("-: ") for c in celdas):   # separadora
                tabla.append(celdas)
            continue
        if tabla:
            cerrar()
        m = _TITULO.match(cruda)
        if m:
            cerrar()
            n = min(len(m.group(1)) + 1, 4)
            partes.append(f"<h{n}>{_inline(m.group(2))}</h{n}>")
            continue
        m = _LISTA.match(cruda)
        if m:
            if not en_lista:
                partes.append("<ul>")
                en_lista = True
            partes.append(f"<li>{_inline(m.group(1))}</li>")
            continue
        cerrar()
        partes.append(f"<p>{_inline(cruda)}</p>")

    cerrar()
    return "\n".join(partes)


def _tabla_html(filas: list[list[str]]) -> str:
    if not filas:
        return ""
    cab, resto = filas[0], filas[1:]
    out = ["<table cellspacing='0'><tr>"]
    out += [f"<th>{_inline(c)}</th>" for c in cab]
    out.append("</tr>")
    for fila in resto:
        out.append("<tr>" + "".join(f"<td>{_inline(c)}</td>" for c in fila) + "</tr>")
    out.append("</table>")
    return "".join(out)


def _esc(s: str) -> str:
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def _inline(s: str) -> str:
    s = _esc(s)
    s = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", s)
    s = re.sub(r"`(.+?)`", r"<code>\1</code>", s)
    return s
